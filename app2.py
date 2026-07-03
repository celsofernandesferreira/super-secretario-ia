import streamlit as st
import google.generativeai as genai
import json
import re
import io
import docx
import logging
from datetime import datetime
import openpyxl  # Biblioteca crucial para abrir e editar o Excel existente

# 1. CONFIGURAÇÃO DE LOGS
logging.basicConfig(
    filename="automacao_rh_raquel.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    encoding="utf-8"
)

st.set_page_config(page_title="Automação de Relatórios RH", page_icon="📊", layout="wide")
st.title("📊 Sistema de Automação de Relatórios — RH")
st.subheader("Extração de Notas para Excel Modelo Predefinido")

# 2. INICIALIZAÇÃO DA API DO GEMINI
try:
    genai.configure(api_key=st.secrets["GOOGLE_API_KEY"])
except Exception:
    st.error("Erro: Chave API em falta nos Secrets do Streamlit.")
    st.stop()

# 3. LEITURA DE TEXTOS (Word e TXT)
def ler_ficheiro_txt(uploaded_file):
    return uploaded_file.read().decode("utf-8")

def ler_ficheiro_docx(uploaded_file):
    doc = docx.Document(io.BytesIO(uploaded_file.read()))
    return "\n".join([para.text for para in doc.paragraphs])

# 4. AGENTE DE INTELIGÊNCIA ARTIFICIAL
def extrair_dados_com_gemini(texto_notas, mapeamento_campos):
    model = genai.GenerativeModel(
        model_name="gemini-3.5-flash",
        generation_config={"response_mime_type": "application/json"}
    )
    
    prompt_sistema = f"""
    Tu és um Assistente de RH Avançado. O teu objetivo é ler as notas fornecidas e extrair os dados necessários para preencher uma tabela corporativa predefinida.
    
    Deves mapear as informações estritamente para as seguintes chaves/colunas que o utilizador vai fornecer:
    {mapeamento_campos}
    
    Regras:
    1. Retorna um array JSON de objetos. Cada objeto é uma linha a introduzir no Excel.
    2. Usa exatamente o nome dos campos fornecidos como chaves do JSON.
    3. Se a informação não existir nas notas, deixa o valor vazio "".
    """
    
    try:
        response = model.generate_content([prompt_sistema, f"Notas para analisar:\n\n{texto_notas}"])
        json_clean = re.sub(r"```json|```", "", response.text).strip()
        return json.loads(json_clean)
    except Exception as e:
        logging.error(f"Erro na extração do Agente: {e}")
        return None

# 5. INTERFACE DO UTILIZADOR
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown("### 📄 1. Documentos Base (Word/Notepad)")
    ficheiros_carregados = st.file_uploader("Carrega as notas de RH", type=["txt", "docx"], accept_multiple_files=True)
    
    texto_acumulado = ""
    if ficheiros_carregados:
        for f in ficheiros_carregados:
            if f.name.endswith(".txt"):
                texto_acumulado += ler_ficheiro_txt(f) + "\n"
            elif f.name.endswith(".docx"):
                texto_acumulado += ler_ficheiro_docx(f) + "\n"
                
    texto_colado = st.text_area("Ou escreve/cola texto adicional aqui:")
    texto_final = (texto_acumulado + "\n" + texto_colado).strip()

with col2:
    st.markdown("### 📂 2. Configuração do Modelo da Empresa")
    
    # Upload do ficheiro Excel oficial da empresa para servir de template
    excel_modelo = st.file_uploader("Carrega aqui o Excel Modelo/Template da Empresa (.xlsx)", type=["xlsx"])
    
    st.caption("Identificação das Colunas (Escreve os nomes exatamente como aparecem no cabeçalho do Excel original, separados por vírgulas):")
    campos_predefinidos = "Nome, Data, Ocorrência, Detalhes"
    campos_usuario = st.text_input("Colunas do Excel original:", value=campos_predefinidos)

st.divider()

# 6. PROCESSAMENTO E INJEÇÃO NO TEMPLATE
if st.button("🤖 Analisar Notas e Preencher Excel da Empresa", use_container_width=True, type="primary"):
    if not texto_final:
        st.warning("Introduz ou carrega notas primeiro.")
    elif not excel_modelo:
        st.error("Precisas de carregar o Excel Modelo da empresa para podermos preenchê-lo.")
    else:
        with st.spinner("O Agente está a ler o documento e a mapear as células no teu Modelo..."):
            
            # O Agente extrai a informação interpretando o texto livre
            lista_dados = extrair_dados_com_gemini(texto_final, campos_usuario)
            
            if lista_dados:
                try:
                    # Carregar o livro de Excel carregado em memória (preserva estilos e fórmulas)
                    wb = openpyxl.load_workbook(io.BytesIO(excel_modelo.read()))
                    ws = wb.active # Escolhe a primeira folha (pode ser alterado para um nome específico)
                    
                    # Transformar a string de campos numa lista limpa
                    colunas_alvo = [c.strip() for c in campos_usuario.split(",")]
                    
                    # 1. Encontrar em que linha estão os cabeçalhos para saber onde escrever
                    linha_cabecalho = 1
                    mapa_colunas_index = {}
                    
                    # Procura nas primeiras 20 linhas o cabeçalho correto
                    for r in range(1, 21):
                        valores_linha = [str(ws.cell(row=r, column=c).value).strip() for c in range(1, ws.max_column + 1)]
                        # Vê se pelo menos um dos campos pedidos está nesta linha
                        if any(campo in valores_linha for campo in colunas_alvo):
                            linha_cabecalho = r
                            # Mapear qual o número da coluna (A=1, B=2...) para cada nome de campo
                            for c in range(1, ws.max_column + 1):
                                nome_celula = str(ws.cell(row=r, column=c).value).strip()
                                if nome_celula in colunas_alvo:
                                    mapa_colunas_index[nome_celula] = c
                            break
                    
                    # Se não encontrou de forma automática, assume que estão na linha 1
                    if not mapa_colunas_index:
                        for idx, nome_col in enumerate(colunas_alvo, start=1):
                            mapa_colunas_index[nome_col] = idx
                    
                    # 2. Descobrir a próxima linha inteiramente vazia para começar a escrever dados
                    proxima_linha = linha_cabecalho + 1
                    while any(ws.cell(row=proxima_linha, column=c).value is not None for c in range(1, ws.max_column + 1)):
                        proxima_linha += 1
                    
                    # 3. Inserir os dados linha a linha respeitando as colunas originais do ficheiro da Raquel
                    linhas_inseridas = 0
                    for registo in lista_dados:
                        for nome_campo, num_coluna in mapa_colunas_index.items():
                            # Retira o valor encontrado pelo Gemini para esta coluna
                            valor_final = registo.get(nome_campo, "")
                            # Grava na célula mantendo o formato original que o Excel já tinha
                            ws.cell(row=proxima_linha, column=num_coluna, value=valor_final)
                        proxima_linha += 1
                        linhas_inseridas += 1
                    
                    # Guardar as alterações num buffer temporário
                    output_final = io.BytesIO()
                    wb.save(output_final)
                    output_final.seek(0)
                    
                    st.success(f"✨ Sucesso! Adicionámos {linhas_inseridas} novas linhas ao documento original da empresa sem alterar a estrutura existente.")
                    
                    # Botão para sacar o ficheiro preenchido
                    st.download_button(
                        label="📥 Descarregar Excel Preenchido",
                        data=output_final,
                        file_name=f"Relatorio_Final_{datetime.now().strftime('%d-%m-%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                except Exception as ex:
                    st.error(f"Erro ao manipular o ficheiro Excel: {ex}")
                    logging.error(f"Erro na manipulação de openpyxl: {ex}")
                    
